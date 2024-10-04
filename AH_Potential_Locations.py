from flask import Flask, request, redirect, url_for, render_template_string, Response
from openpyxl import load_workbook
from io import BytesIO
from datetime import datetime

app = Flask(__name__)

@app.route('/')
def upload_file():
    return render_template_string('''
    <!doctype html>
    <html lang="es">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Subir archivo Excel - AH Potential Locations</title>
        <link href="https://fonts.googleapis.com/css2?family=Roboto:wght@400;700&display=swap" rel="stylesheet">
        <style>
            body {
                font-family: 'Roboto', sans-serif;
                background-color: #f8f9fa;
                margin: 0;
                padding: 0;
                display: flex;
                justify-content: center;
                align-items: center;
                height: 100vh;
            }
            .container {
                background: #ffffff;
                padding: 30px;
                box-shadow: 0 6px 15px rgba(0, 0, 0, 0.1);
                border-radius: 10px;
                text-align: center;
                max-width: 90%;
                width: 400px;
            }
            h1 {
                color: #333;
                font-weight: 700;
                margin-bottom: 20px;
            }
            form {
                margin-top: 20px;
            }
            input[type=file] {
                margin-bottom: 20px;
            }
            input[type=submit] {
                background-color: #ff4c00;
                color: #ffffff;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                cursor: pointer;
                font-weight: bold;
                transition: background-color 0.3s ease;
            }
            input[type=submit]:hover {
                background-color: #e04300;
            }
        </style>
    </head>
    <body>
        <div class="container">
            <h1>Sube el archivo Excel</h1>
            <form method="post" enctype="multipart/form-data" action="/upload">
                <input type="file" name="file" required>
                <br>
                <input type="submit" value="Subir archivo">
            </form>
        </div>
    </body>
    </html>
    ''')

@app.route('/upload', methods=['POST'])
def upload():
    if 'file' not in request.files:
        return 'No file part'
    file = request.files['file']
    if file.filename == '':
        return 'No selected file'
    if file:
        wb = load_workbook(filename=BytesIO(file.read()), data_only=True)
        global uploaded_wb
        uploaded_wb = wb
        return redirect(url_for('display_data'))

@app.route('/data', methods=['GET'])
def display_data():
    if 'uploaded_wb' not in globals():
        return redirect(url_for('upload_file'))

    wb = uploaded_wb
    sheet = wb.active

    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        location = row[9]
        google_map = f'<a href="{row[14]}" target="_blank">Link</a>' if row[14] else 'Missing'
        pictures = f'<a href="{row[15]}" target="_blank">Link</a>' if row[15] else 'Missing'
        real_estate_ad = f'<a href="{row[16]}" target="_blank">Link</a>' if row[16] else 'Missing'
        net_rent_month = row[21]
        total_sqm = row[27]
        estimated_parking_spots = row[41]
        rent_per_sqm = row[39]
        cost_per_parking_spot = row[1]
        flexicar = f'<a href="{row[3]}" target="_blank">Link</a>' if row[3] else 'Missing'
        ocasionplus = f'<a href="{row[4]}" target="_blank">Link</a>' if row[4] else 'Missing'
        ctc = f'<a href="{row[5]}" target="_blank">Link</a>' if row[5] else 'Missing'
        comments = row[40] if row[40] else 'No comments available'

        data.append((
            location, google_map, pictures, real_estate_ad,
            net_rent_month, total_sqm, estimated_parking_spots,
            rent_per_sqm, cost_per_parking_spot,
            flexicar, ocasionplus, ctc, comments
        ))

    location_filter = request.args.get('location_filter', '')
    if location_filter:
        data = [row for row in data if row[0] and location_filter.lower() in row[0].lower()]

    html_table = '''
    <style>
        body {
            font-family: 'Roboto', sans-serif;
            background-color: #f8f9fa;
            padding: 20px;
        }
        .container {
            max-width: 1200px;
            margin: auto;
            background: #ffffff;
            padding: 20px;
            box-shadow: 0 6px 15px rgba(0, 0, 0, 0.1);
            border-radius: 10px;
        }
        table {
            width: 100%;
            border-collapse: collapse;
            margin-top: 20px;
        }
        th, td {
            padding: 15px;
            text-align: center;
            border-bottom: 1px solid #ddd;
            font-size: 0.95em;
        }
        th {
            background-color: #ff4c00;
            color: white;
            text-transform: uppercase;
        }
        tr:hover {
            background-color: #f1f1f1;
        }
        a {
            color: #ff4c00;
            text-decoration: none;
            font-weight: bold;
        }
        a:hover {
            text-decoration: underline;
        }
        form {
            display: flex;
            gap: 10px;
            margin-bottom: 20px;
        }
        input[type=submit] {
            background-color: #ff4c00;
            color: #ffffff;
            padding: 10px;
            border: none;
            border-radius: 5px;
            cursor: pointer;
            font-weight: bold;
            transition: background-color 0.3s ease;
        }
        input[type=submit]:hover {
            background-color: #e04300;
        }
        .expandable-row {
            display: none;
            background-color: #f9f9f9;
        }
        .expanded-content {
            padding: 15px;
            text-align: left;
            color: #333;
        }
        .clickable-info {
            color: #ff4c00;
            font-weight: bold;
            cursor: pointer;
        }
    </style>
    <script>
        function toggleRow(id) {
            var row = document.getElementById(id);
            if (row.style.display === "none" || row.style.display === "") {
                row.style.display = "table-row";
            } else {
                row.style.display = "none";
            }
        }
    </script>
    <div class="container">
        <form method="get">
            <label for="location_filter">Filtrar por ubicación:</label>
            <input type="text" name="location_filter" id="location_filter" placeholder="Ubicación">
            <input type="submit" value="Aplicar">
        </form>
        <table>
            <tr>
                <th>#</th>
                <th>Ubicación</th>
                <th>Google Map</th>
                <th>Imágenes</th>
                <th>Anuncio Inmobiliario</th>
                <th>Renta neta / mes</th>
                <th>Superficie total (m²)</th>
                <th>Plazas de aparcamiento estimadas</th>
                <th>Renta por m²</th>
                <th>€/Plaza de aparcamiento</th>
                <th>+Info</th>
            </tr>
    '''

    for index, row in enumerate(data):
        expandable_row_id = f"expandable-row-{index}"
        html_table += f'''
            <tr>
                <td>{index + 1}</td>
                <td>{row[0]}</td>
                <td>{row[1]}</td>
                <td>{row[2]}</td>
                <td>{row[3]}</td>
                <td>{row[4]}</td>
                <td>{row[5]}</td>
                <td>{row[6]}</td>
                <td>{row[7]}</td>
                <td>{row[8]}</td>
                <td class="clickable-info" onclick="toggleRow('{expandable_row_id}')">+info</td>
            </tr>
            <tr id="{expandable_row_id}" class="expandable-row">
                <td colspan="11" class="expanded-content">
                    <p><strong>Flexicar:</strong> {row[9]}</p>
                    <p><strong>OcasionPlus:</strong> {row[10]}</p>
                    <p><strong>CTC:</strong> {row[11]}</p>
                    <p><strong>Comentarios:</strong> {row[12]}</p>
                </td>
            </tr>
        '''

    html_table += '''
        </table>
        <form method="post" action="/generate_report">
            <button type="submit" class="generate-report-button">Generar Reporte</button>
        </form>
    </div>
    '''

    current_date = datetime.now().strftime("%d/%m/%Y")

    return render_template_string('''
    <!doctype html>
    <html lang="es">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Report AH Potential Locations</title>
        <link href="https://fonts.googleapis.com/css2?family=Roboto:wght@400;700&display=swap" rel="stylesheet">
    </head>
    <body>
        <div class="container">
            <h1>Reporte - AH Potential Locations</h1>
            <h2>Fecha: {{ date }}</h2>
            <div>{{ table|safe }}</div>
        </div>
    </body>
    </html>
    ''', table=html_table, date=current_date)

@app.route('/generate_report', methods=['POST'])
def generate_report():
    if 'uploaded_wb' not in globals():
        return redirect(url_for('upload_file'))

    wb = uploaded_wb
    sheet = wb.active

    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        location = row[9]
        google_map = f'<a href="{row[14]}" target="_blank">Link</a>' if row[14] else 'Missing'
        pictures = f'<a href="{row[15]}" target="_blank">Link</a>' if row[15] else 'Missing'
        real_estate_ad = f'<a href="{row[16]}" target="_blank">Link</a>' if row[16] else 'Missing'
        net_rent_month = row[21]
        total_sqm = row[27]
        estimated_parking_spots = row[41]
        rent_per_sqm = row[39]
        cost_per_parking_spot = row[1]
        flexicar = f'<a href="{row[3]}" target="_blank">Link</a>' if row[3] else 'Missing'
        ocasionplus = f'<a href="{row[4]}" target="_blank">Link</a>' if row[4] else 'Missing'
        ctc = f'<a href="{row[5]}" target="_blank">Link</a>' if row[5] else 'Missing'
        comments = row[40] if row[40] else 'No comments available'

        data.append((
            location, google_map, pictures, real_estate_ad,
            net_rent_month, total_sqm, estimated_parking_spots,
            rent_per_sqm, cost_per_parking_spot,
            flexicar, ocasionplus, ctc, comments
        ))

    html_content = render_template_string('''
    <!doctype html>
    <html lang="es">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Generated Report - AH Potential Locations</title>
        <link href="https://fonts.googleapis.com/css2?family=Roboto:wght@400;700&display=swap" rel="stylesheet">
        <style>
            body {
                font-family: 'Roboto', sans-serif;
                background-color: #f8f9fa;
                padding: 20px;
            }
            .container {
                max-width: 1200px;
                margin: auto;
                background: #ffffff;
                padding: 20px;
                box-shadow: 0 6px 15px rgba(0, 0, 0, 0.1);
                border-radius: 10px;
            }
            table {
                width: 100%;
                border-collapse: collapse;
                margin-top: 20px;
            }
            th, td {
                padding: 15px;
                text-align: center;
                border-bottom: 1px solid #ddd;
                font-size: 0.95em;
            }
            th {
                background-color: #ff4c00;
                color: white;
                text-transform: uppercase;
            }
            tr:hover {
                background-color: #f1f1f1;
            }
            a {
                color: #ff4c00;
                text-decoration: none;
                font-weight: bold;
            }
            a:hover {
                text-decoration: underline;
            }
        </style>
    </head>
    <body>
        <div class="container">
            <h1>Generated Report - AH Potential Locations</h1>
            <h2>Fecha: {{ date }}</h2>
            <table>
                <tr>
                    <th>#</th>
                    <th>Ubicación</th>
                    <th>Google Map</th>
                    <th>Imágenes</th>
                    <th>Anuncio Inmobiliario</th>
                    <th>Renta neta / mes</th>
                    <th>Superficie total (m²)</th>
                    <th>Plazas de aparcamiento estimadas</th>
                    <th>Renta por m²</th>
                    <th>€/Plaza de aparcamiento</th>
                    <th>+Info</th>
                </tr>
                {% for index, row in data %}
                <tr>
                    <td>{{ index + 1 }}</td>
                    <td>{{ row[0] }}</td>
                    <td>{{ row[1]|safe }}</td>
                    <td>{{ row[2]|safe }}</td>
                    <td>{{ row[3]|safe }}</td>
                    <td>{{ row[4] }}</td>
                    <td>{{ row[5] }}</td>
                    <td>{{ row[6] }}</td>
                    <td>{{ row[7] }}</td>
                    <td>{{ row[8] }}</td>
                    <td>
                        <p><strong>Flexicar:</strong> {{ row[9]|safe }}</p>
                        <p><strong>OcasionPlus:</strong> {{ row[10]|safe }}</p>
                        <p><strong>CTC:</strong> {{ row[11]|safe }}</p>
                        <p><strong>Comentarios:</strong> {{ row[12] }}</p>
                    </td>
                </tr>
                {% endfor %}
            </table>
        </div>
    </body>
    </html>
    ''', data=enumerate(data), date=datetime.now().strftime("%d/%m/%Y"))

    response = Response(html_content)
    response.headers['Content-Disposition'] = 'attachment; filename=report.html'
    response.mimetype = 'text/html'
    return response

# Configuración para Vercel
app = app

if __name__ == '__main__':
    app.run()



